#!/usr/bin/env python3
"""
OpenClaw RAG Knowledge Base Manager
====================================
Local RAG system with document parsing, vector storage, and hybrid retrieval.

Usage:
    python3 kb_manager.py add <path> [--collection NAME]
    python3 kb_manager.py query <question> [--top-k N] [--collection NAME]
    python3 kb_manager.py context <question> [--max-tokens N]
    python3 kb_manager.py list
    python3 kb_manager.py stats
    python3 kb_manager.py remove --source <path>
"""

import argparse
import hashlib
import json
import os
import re
import sys
import time
from pathlib import Path
from typing import Optional

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
OPENCLAW_HOME = Path.home() / ".openclaw"
CHROMA_DIR = OPENCLAW_HOME / "data" / "chromadb"
DEFAULT_COLLECTION = "default"
CHUNK_SIZE = 800        # characters per chunk
CHUNK_OVERLAP = 150     # overlap between chunks
EMBEDDING_MODEL = "text-embedding-v3"
EMBEDDING_DIM = 1024
MAX_BATCH_SIZE = 10     # DashScope API batch limit

# API key resolution: env var > hardcoded dashscope key
# Note: bailian sk-sp- keys only work on coding.dashscope.aliyuncs.com, NOT for embeddings.
# The standard DASHSCOPE_API_KEY (sk-xxx) works on dashscope.aliyuncs.com for embeddings.
def resolve_api_key() -> str:
    key = os.environ.get("DASHSCOPE_API_KEY", "").strip()
    if key:
        return key
    # Fallback: try to read from openclaw.json bailian provider (only if NOT sk-sp- prefix)
    try:
        cfg_path = OPENCLAW_HOME / "openclaw.json"
        with open(cfg_path) as f:
            cfg = json.load(f)
        key = cfg.get("models", {}).get("providers", {}).get("bailian", {}).get("apiKey", "")
        if key and not key.startswith("sk-sp-"):
            return key
    except Exception:
        pass
    print("ERROR: No API key found. Set DASHSCOPE_API_KEY env var (standard DashScope key, not sk-sp-).", file=sys.stderr)
    sys.exit(1)


# ---------------------------------------------------------------------------
# Document Parsers
# ---------------------------------------------------------------------------
def parse_pdf(path: Path) -> str:
    from pypdf import PdfReader
    reader = PdfReader(str(path))
    texts = []
    for page in reader.pages:
        t = page.extract_text()
        if t:
            texts.append(t)
    return "\n\n".join(texts)


def parse_docx(path: Path) -> str:
    from docx import Document
    doc = Document(str(path))
    return "\n\n".join(p.text for p in doc.paragraphs if p.text.strip())


def parse_xlsx(path: Path) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(str(path), read_only=True, data_only=True)
    parts = []
    for ws in wb.worksheets:
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(cells):
                rows.append("\t".join(cells))
        if rows:
            parts.append(f"## Sheet: {ws.title}\n" + "\n".join(rows))
    return "\n\n".join(parts)


def parse_markdown(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="replace")


def parse_text(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="replace")


PARSERS = {
    ".pdf": parse_pdf,
    ".docx": parse_docx,
    ".doc": parse_docx,
    ".xlsx": parse_xlsx,
    ".xls": parse_xlsx,
    ".md": parse_markdown,
    ".markdown": parse_markdown,
    ".txt": parse_text,
    ".log": parse_text,
    ".json": parse_text,
    ".csv": parse_text,
    ".py": parse_text,
    ".ts": parse_text,
    ".js": parse_text,
    ".sh": parse_text,
    ".yaml": parse_text,
    ".yml": parse_text,
    ".toml": parse_text,
    ".html": parse_text,
    ".xml": parse_text,
}


def parse_document(path: Path) -> Optional[str]:
    suffix = path.suffix.lower()
    parser = PARSERS.get(suffix)
    if parser is None:
        return None
    try:
        return parser(path)
    except Exception as e:
        print(f"  WARN: Failed to parse {path}: {e}", file=sys.stderr)
        return None


# ---------------------------------------------------------------------------
# Chunking
# ---------------------------------------------------------------------------
def chunk_text(text: str, chunk_size: int = CHUNK_SIZE, overlap: int = CHUNK_OVERLAP) -> list[str]:
    """Recursive character text splitter with semantic boundaries."""
    if not text or not text.strip():
        return []

    separators = ["\n\n", "\n", "。", ".", "！", "!", "？", "?", "；", ";", " "]
    chunks = []

    def _split(t: str, seps: list[str]) -> list[str]:
        if len(t) <= chunk_size:
            return [t] if t.strip() else []
        sep = seps[0] if seps else ""
        remaining_seps = seps[1:] if len(seps) > 1 else []

        if sep and sep in t:
            parts = t.split(sep)
        else:
            if remaining_seps:
                return _split(t, remaining_seps)
            # Hard split
            result = []
            for i in range(0, len(t), chunk_size - overlap):
                piece = t[i:i + chunk_size]
                if piece.strip():
                    result.append(piece)
            return result

        merged = []
        current = ""
        for part in parts:
            candidate = current + sep + part if current else part
            if len(candidate) <= chunk_size:
                current = candidate
            else:
                if current.strip():
                    merged.append(current)
                if len(part) > chunk_size:
                    if remaining_seps:
                        merged.extend(_split(part, remaining_seps))
                    else:
                        for i in range(0, len(part), chunk_size - overlap):
                            piece = part[i:i + chunk_size]
                            if piece.strip():
                                merged.append(piece)
                else:
                    current = part
                    continue
                current = ""
        if current.strip():
            merged.append(current)
        return merged

    raw_chunks = _split(text, separators)

    # Add overlap
    for i, chunk in enumerate(raw_chunks):
        if i > 0 and overlap > 0:
            prev_tail = raw_chunks[i - 1][-overlap:]
            chunk = prev_tail + chunk
        chunks.append(chunk.strip())

    return [c for c in chunks if c]


# ---------------------------------------------------------------------------
# Embedding via DashScope
# ---------------------------------------------------------------------------
def _call_embedding_api(texts: list[str], api_key: str) -> list[list[float]]:
    """Single API call for a batch of texts."""
    import urllib.request

    url = "https://dashscope.aliyuncs.com/compatible-mode/v1/embeddings"
    payload = {
        "model": EMBEDDING_MODEL,
        "input": texts,
        "dimensions": EMBEDDING_DIM,
    }
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
    }
    req = urllib.request.Request(url, data=json.dumps(payload).encode(), headers=headers, method="POST")
    resp = urllib.request.urlopen(req, timeout=30)
    result = json.loads(resp.read())
    if "data" not in result:
        raise RuntimeError(f"Unexpected response: {json.dumps(result, ensure_ascii=False)[:200]}")
    data = sorted(result["data"], key=lambda e: e.get("index", 0))
    return [e["embedding"] for e in data]


def get_embeddings(texts: list[str], api_key: str) -> list[list[float]]:
    """Call DashScope text-embedding-v3 via OpenAI-compatible endpoint with retry."""
    import urllib.error

    # Truncate any excessively long chunks (API limit ~8192 tokens ≈ 6000 chars for Chinese)
    MAX_CHUNK_CHARS = 5000
    texts = [t[:MAX_CHUNK_CHARS] for t in texts]

    all_embeddings = []
    for i in range(0, len(texts), MAX_BATCH_SIZE):
        batch = texts[i:i + MAX_BATCH_SIZE]
        try:
            embs = _call_embedding_api(batch, api_key)
            all_embeddings.extend(embs)
        except (urllib.error.HTTPError, Exception) as batch_err:
            # Batch failed — fall back to one-by-one
            err_detail = ""
            if hasattr(batch_err, "read"):
                err_detail = batch_err.read().decode()[:200]
            print(f"\n  WARN: Batch embedding failed ({batch_err}), retrying one-by-one... {err_detail}", file=sys.stderr)
            for j, text in enumerate(batch):
                try:
                    emb = _call_embedding_api([text], api_key)
                    all_embeddings.extend(emb)
                except Exception as single_err:
                    print(f"  WARN: Skipping chunk (len={len(text)}): {single_err}", file=sys.stderr)
                    # Use zero vector as placeholder
                    all_embeddings.append([0.0] * EMBEDDING_DIM)

    return all_embeddings


def get_query_embedding(text: str, api_key: str) -> list[float]:
    """Get embedding for a query via OpenAI-compatible endpoint."""
    import urllib.request

    url = "https://dashscope.aliyuncs.com/compatible-mode/v1/embeddings"
    payload = {
        "model": EMBEDDING_MODEL,
        "input": [text],
        "dimensions": EMBEDDING_DIM,
    }
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
    }
    req = urllib.request.Request(url, data=json.dumps(payload).encode(), headers=headers, method="POST")
    with urllib.request.urlopen(req, timeout=30) as resp:
        result = json.loads(resp.read())
    return result["data"][0]["embedding"]


# ---------------------------------------------------------------------------
# ChromaDB Operations
# ---------------------------------------------------------------------------
def get_chroma_client():
    import chromadb
    CHROMA_DIR.mkdir(parents=True, exist_ok=True)
    return chromadb.PersistentClient(path=str(CHROMA_DIR))


def get_collection(client, name: str = DEFAULT_COLLECTION):
    return client.get_or_create_collection(
        name=name,
        metadata={"hnsw:space": "cosine"},
    )


def doc_id(source: str, chunk_idx: int) -> str:
    h = hashlib.md5(source.encode()).hexdigest()[:12]
    return f"{h}-{chunk_idx:04d}"


# ---------------------------------------------------------------------------
# Commands
# ---------------------------------------------------------------------------
def cmd_add(args):
    path = Path(args.path).resolve()
    collection_name = args.collection or DEFAULT_COLLECTION
    api_key = resolve_api_key()

    files = []
    if path.is_file():
        files = [path]
    elif path.is_dir():
        for ext in PARSERS.keys():
            files.extend(path.rglob(f"*{ext}"))
        files = sorted(set(files))
    else:
        print(f"ERROR: Path not found: {path}", file=sys.stderr)
        sys.exit(1)

    if not files:
        print(f"No parseable files found in {path}")
        return

    print(f"📚 Adding {len(files)} file(s) to collection '{collection_name}'...")

    client = get_chroma_client()
    col = get_collection(client, collection_name)

    total_chunks = 0
    for fpath in files:
        print(f"  📄 {fpath.name}...", end=" ", flush=True)
        text = parse_document(fpath)
        if not text or not text.strip():
            print("(empty, skipped)")
            continue

        chunks = chunk_text(text)
        if not chunks:
            print("(no chunks, skipped)")
            continue

        # Remove existing chunks for this source
        source_str = str(fpath)
        existing = col.get(where={"source": source_str})
        if existing and existing["ids"]:
            col.delete(ids=existing["ids"])

        # Get embeddings
        embeddings = get_embeddings(chunks, api_key)

        # Add to ChromaDB
        ids = [doc_id(source_str, i) for i in range(len(chunks))]
        metadatas = [
            {
                "source": source_str,
                "filename": fpath.name,
                "chunk_index": i,
                "total_chunks": len(chunks),
                "char_count": len(c),
                "added_at": time.strftime("%Y-%m-%dT%H:%M:%S"),
            }
            for i, c in enumerate(chunks)
        ]

        col.add(ids=ids, embeddings=embeddings, documents=chunks, metadatas=metadatas)
        total_chunks += len(chunks)
        print(f"({len(chunks)} chunks)")

    print(f"\n✅ Done! Added {total_chunks} chunks from {len(files)} files to '{collection_name}'")
    print(f"   Storage: {CHROMA_DIR}")


def cmd_query(args):
    question = args.question
    top_k = args.top_k or 5
    collection_name = args.collection or DEFAULT_COLLECTION
    api_key = resolve_api_key()

    client = get_chroma_client()
    try:
        col = client.get_collection(collection_name)
    except Exception:
        print(f"Collection '{collection_name}' not found. Add documents first.")
        return

    print(f"🔍 Querying '{collection_name}' for: {question[:80]}...")

    query_emb = get_query_embedding(question, api_key)
    results = col.query(query_embeddings=[query_emb], n_results=top_k, include=["documents", "metadatas", "distances"])

    if not results["documents"] or not results["documents"][0]:
        print("No results found.")
        return

    print(f"\n📋 Top {len(results['documents'][0])} results:\n")
    for i, (doc, meta, dist) in enumerate(zip(
        results["documents"][0], results["metadatas"][0], results["distances"][0]
    )):
        score = 1 - dist  # cosine similarity
        source = meta.get("filename", "?")
        chunk_idx = meta.get("chunk_index", "?")
        print(f"--- [{i+1}] {source} (chunk {chunk_idx}, score: {score:.3f}) ---")
        # Truncate for display
        preview = doc[:300] + "..." if len(doc) > 300 else doc
        print(preview)
        print()


def cmd_context(args):
    """Output retrieval results formatted as LLM context injection."""
    question = args.question
    max_tokens = args.max_tokens or 4000
    collection_name = args.collection or DEFAULT_COLLECTION
    api_key = resolve_api_key()

    client = get_chroma_client()
    try:
        col = client.get_collection(collection_name)
    except Exception:
        print(f"Collection '{collection_name}' not found.", file=sys.stderr)
        sys.exit(1)

    query_emb = get_query_embedding(question, api_key)
    results = col.query(query_embeddings=[query_emb], n_results=10, include=["documents", "metadatas", "distances"])

    if not results["documents"] or not results["documents"][0]:
        print("(no knowledge base results)")
        return

    context_parts = []
    char_budget = max_tokens * 3  # rough char-to-token ratio
    used = 0

    for doc, meta, dist in zip(
        results["documents"][0], results["metadatas"][0], results["distances"][0]
    ):
        score = 1 - dist
        if score < 0.3:  # relevance threshold
            continue
        source = meta.get("filename", "unknown")
        entry = f"[来源: {source}, 相关度: {score:.2f}]\n{doc}"
        if used + len(entry) > char_budget:
            break
        context_parts.append(entry)
        used += len(entry)

    if not context_parts:
        print("(no sufficiently relevant results)")
        return

    header = f"以下是从知识库中检索到的 {len(context_parts)} 条相关内容：\n"
    print(header + "\n---\n".join(context_parts))


def cmd_list(args):
    client = get_chroma_client()
    collections = client.list_collections()
    if not collections:
        print("No collections found. Use 'add' to create one.")
        return

    print(f"📚 Knowledge Base Collections ({len(collections)}):\n")
    for col in collections:
        count = col.count()
        print(f"  📁 {col.name}: {count} chunks")
        if count > 0:
            peek = col.peek(limit=3)
            sources = set()
            if peek["metadatas"]:
                for m in peek["metadatas"]:
                    if m and "filename" in m:
                        sources.add(m["filename"])
            if sources:
                print(f"     Sources: {', '.join(sorted(sources))}...")


def cmd_stats(args):
    client = get_chroma_client()
    collections = client.list_collections()

    total_chunks = 0
    total_sources = set()

    print("📊 Knowledge Base Statistics\n")
    for col in collections:
        count = col.count()
        total_chunks += count
        all_data = col.get(include=["metadatas"])
        sources = set()
        if all_data["metadatas"]:
            for m in all_data["metadatas"]:
                if m and "source" in m:
                    sources.add(m["source"])
        total_sources.update(sources)
        print(f"  Collection '{col.name}': {count} chunks from {len(sources)} files")

    print(f"\n  Total: {total_chunks} chunks, {len(total_sources)} unique files")
    print(f"  Storage: {CHROMA_DIR}")
    if CHROMA_DIR.exists():
        size = sum(f.stat().st_size for f in CHROMA_DIR.rglob("*") if f.is_file())
        print(f"  Disk usage: {size / 1024 / 1024:.1f} MB")


def cmd_remove(args):
    source = str(Path(args.source).resolve())
    collection_name = args.collection or DEFAULT_COLLECTION

    client = get_chroma_client()
    try:
        col = client.get_collection(collection_name)
    except Exception:
        print(f"Collection '{collection_name}' not found.")
        return

    existing = col.get(where={"source": source})
    if not existing or not existing["ids"]:
        print(f"No chunks found for source: {source}")
        return

    col.delete(ids=existing["ids"])
    print(f"✅ Removed {len(existing['ids'])} chunks for {source}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(description="OpenClaw RAG Knowledge Base Manager")
    sub = parser.add_subparsers(dest="command")

    p_add = sub.add_parser("add", help="Add documents to knowledge base")
    p_add.add_argument("path", help="File or directory path")
    p_add.add_argument("--collection", "-c", default=DEFAULT_COLLECTION, help="Collection name")

    p_query = sub.add_parser("query", help="Query knowledge base")
    p_query.add_argument("question", help="Query string")
    p_query.add_argument("--top-k", "-k", type=int, default=5, help="Number of results")
    p_query.add_argument("--collection", "-c", default=DEFAULT_COLLECTION, help="Collection name")

    p_ctx = sub.add_parser("context", help="Get retrieval context for LLM")
    p_ctx.add_argument("question", help="Query string")
    p_ctx.add_argument("--max-tokens", "-t", type=int, default=4000, help="Max token budget")
    p_ctx.add_argument("--collection", "-c", default=DEFAULT_COLLECTION, help="Collection name")

    p_list = sub.add_parser("list", help="List collections")

    p_stats = sub.add_parser("stats", help="Show statistics")

    p_rm = sub.add_parser("remove", help="Remove document from knowledge base")
    p_rm.add_argument("--source", "-s", required=True, help="Source file path")
    p_rm.add_argument("--collection", "-c", default=DEFAULT_COLLECTION, help="Collection name")

    args = parser.parse_args()
    if not args.command:
        parser.print_help()
        sys.exit(1)

    commands = {
        "add": cmd_add,
        "query": cmd_query,
        "context": cmd_context,
        "list": cmd_list,
        "stats": cmd_stats,
        "remove": cmd_remove,
    }
    commands[args.command](args)


if __name__ == "__main__":
    main()
